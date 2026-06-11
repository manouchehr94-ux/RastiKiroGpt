"""
Tenants - Admin Views.

CRUD views for company admins to manage:
- Company public page
- Services (create/edit/delete)
- Technicians (create/edit/delete)
- Customers (list/detail)
- Orders (list/filter/assign/cancel)
- Invoices (list/filter/issue/cancel)

All views are thin — business logic in services/selectors.
All data scoped by request.company (tenant isolation).
"""
from django.http import Http404, HttpRequest, HttpResponse, HttpResponseForbidden
from django.shortcuts import redirect, render

from apps.common.jalali import format_jalali_date, normalize_digits, parse_jalali_date, today_jalali_date

from apps.accounts.models import CompanyUser, Customer, Technician, TechnicianCategorySkill, TechnicianSkill, UserRole
from apps.accounts.permissions import require_tenant_role
from apps.invoices.models import Invoice
from apps.invoices.selectors import InvoiceSelector
from apps.invoices.services import (
    InvoiceCancelService,
    InvoiceCreateService,
    InvoiceIssueService,
    InvoiceMarkPaidService,
    InvoiceUpdateService,
)
from apps.orders.models import Order
from apps.orders.selectors import OrderSelector
from apps.orders.services import OrderCancelService

from .forms import (
    CompanyPageForm,
    CompanyServiceForm,
    TechnicianCreateForm,
    TechnicianEditForm,
)
from .models import CompanyService, ServiceRequest
from .selectors import (
    CompanyGallerySelector,
    CompanyPageSelector,
    CompanyServiceSelector,
    ServiceRequestSelector,
)
from .services import (
    CompanyPageUpdateService,
    CompanyServiceCreateService,
    CompanyServiceUpdateService,
)


# =============================================================================
# COMPANY SETTINGS
# =============================================================================


def _parse_time_from_post(raw_value: str):
    """Parse HH:MM or HH:MM:SS into a Python time object."""
    from datetime import time

    raw_value = normalize_digits((raw_value or "").strip())
    if not raw_value:
        return None
    parts = raw_value.split(":")
    if len(parts) not in (2, 3):
        raise ValueError("ساعت باید با فرمت 07:30 یا 07:30:00 وارد شود.")
    try:
        hour = int(parts[0])
        minute = int(parts[1])
        second = int(parts[2]) if len(parts) == 3 else 0
        return time(hour=hour, minute=minute, second=second)
    except ValueError as exc:
        raise ValueError("ساعت وارد شده معتبر نیست.") from exc


def _parse_non_negative_int_from_post(request: HttpRequest, field_name: str, default: int = 0) -> int:
    raw = normalize_digits(request.POST.get(field_name, str(default))).strip()
    if raw == "":
        return default
    try:
        return max(0, int(raw))
    except ValueError as exc:
        raise ValueError(f"مقدار {field_name} باید عدد صحیح باشد.") from exc


@require_tenant_role("COMPANY_ADMIN")
def admin_company_settings(request: HttpRequest, **kwargs) -> HttpResponse:
    """Edit company profile and operational order/technician settings."""
    from apps.tenants.selectors import get_company_settings

    company = request.company
    settings_obj = get_company_settings(company)
    error = ""
    success = ""

    from apps.tenants.models import CompanyFinancialPolicy
    financial_policy, _ = CompanyFinancialPolicy.objects.get_or_create(
        company=company,
        defaults={
            "campaign_discount_policy": CompanyFinancialPolicy.DiscountPolicy.COMPANY,
            "extra_discount_policy": CompanyFinancialPolicy.DiscountPolicy.TECHNICIAN,
            "payout_strategy": CompanyFinancialPolicy.PayoutStrategy.DIRECT_TO_COMPANY,
        },
    )

    if request.method == "POST":
        try:
            company.name = request.POST.get("name", "").strip() or company.name
            company.address = request.POST.get("address", "").strip()
            company.phone = request.POST.get("phone", "").strip()
            company.email = request.POST.get("email", "").strip()
            if hasattr(company, "economic_code"):
                company.economic_code = request.POST.get("economic_code", "").strip()
            if hasattr(company, "website"):
                company.website = request.POST.get("website", "").strip()
            company.save(update_fields=[
                "name", "address", "phone", "email", "economic_code", "website", "updated_at"
            ])

            settings_obj.max_active_orders_per_technician = _parse_non_negative_int_from_post(
                request, "max_active_orders_per_technician", settings_obj.max_active_orders_per_technician
            )
            settings_obj.priority2_delay_minutes = _parse_non_negative_int_from_post(
                request, "priority2_delay_minutes", settings_obj.priority2_delay_minutes
            )
            settings_obj.priority3_delay_minutes = _parse_non_negative_int_from_post(
                request, "priority3_delay_minutes", settings_obj.priority3_delay_minutes
            )
            settings_obj.show_future_orders_to_technicians = bool(
                request.POST.get("show_future_orders_to_technicians")
            )
            settings_obj.future_orders_visible_after = _parse_time_from_post(
                request.POST.get("future_orders_visible_after", "")
            )
            settings_obj.auto_recycle_cancel_request = bool(
                request.POST.get("auto_recycle_cancel_request")
            )
            settings_obj.save(update_fields=[
                "max_active_orders_per_technician",
                "priority2_delay_minutes",
                "priority3_delay_minutes",
                "show_future_orders_to_technicians",
                "future_orders_visible_after",
                "auto_recycle_cancel_request",
            ])

            # Payout strategy — company admin may change strategy; platform_fee_percent is read-only here
            new_strategy = request.POST.get("payout_strategy", "")
            valid_strategies = {c[0] for c in CompanyFinancialPolicy.PayoutStrategy.choices}
            if new_strategy in valid_strategies:
                financial_policy.payout_strategy = new_strategy
            financial_policy.save(update_fields=["payout_strategy", "updated_at"])

            success = "تنظیمات با موفقیت ذخیره شد."
        except ValueError as exc:
            error = str(exc)

    return render(request, "tenants/admin_company_settings.html", {
        "company": company,
        "settings_obj": settings_obj,
        "financial_policy": financial_policy,
        "error": error,
        "success": success,
    })





@require_tenant_role("COMPANY_ADMIN")
def admin_notification_settings(request: HttpRequest, **kwargs) -> HttpResponse:
    """Enable/disable in-app and SMS notifications per company event."""
    from apps.notifications.services import NotificationSettingService

    company = request.company
    error = ""
    success = ""

    if request.method == "POST":
        NotificationSettingService.update_from_post(
            company=company,
            post_data=request.POST,
        )
        success = "تنظیمات نوتیفیکیشن با موفقیت ذخیره شد."

    notification_settings = NotificationSettingService.ensure_defaults(company=company)

    return render(request, "tenants/admin_notification_settings.html", {
        "company": company,
        "notification_settings": notification_settings,
        "error": error,
        "success": success,
    })


# =============================================================================
# COMPANY PAGE
# =============================================================================


@require_tenant_role("COMPANY_ADMIN")
def admin_page_edit(request: HttpRequest, **kwargs) -> HttpResponse:
    """Admin page for editing company public page content."""
    company = request.company
    page = CompanyPageSelector.get_for_company(company=company)

    if request.method == "POST":
        form = CompanyPageForm(request.POST)
        if form.is_valid():
            CompanyPageUpdateService.update(company=company, data=form.cleaned_data)
            return redirect(f"/{company.code}/admin/page/")
    else:
        form = CompanyPageForm(initial={
            "title": page.title,
            "intro_text": page.intro_text,
            "contact_phone": page.contact_phone,
            "contact_email": page.contact_email,
            "address": page.address,
            "working_hours": page.working_hours,
            "is_request_form_enabled": page.is_request_form_enabled,
            "is_published": page.is_published,
        })

    return render(request, "tenants/admin_page_edit.html", {
        "company": company, "page": page, "form": form,
    })


# =============================================================================
# COMPANY SERVICES CRUD
# =============================================================================



def _save_technician_category_priorities(*, technician: Technician, company, post_data):
    """Persist technician ↔ service category priority mapping from admin form."""
    from apps.tenants.models import CompanyServiceCategory

    categories = CompanyServiceCategory.objects.filter(company=company)
    valid_priorities = {"1", "2", "3"}
    for category in categories:
        raw = post_data.get(f"category_priority_{category.id}", "")
        if raw not in valid_priorities:
            TechnicianCategorySkill.objects.filter(
                technician=technician,
                category=category,
            ).delete()
            continue
        TechnicianCategorySkill.objects.update_or_create(
            technician=technician,
            category=category,
            defaults={"priority": int(raw)},
        )


# =============================================================================
# TECHNICIANS CRUD
# =============================================================================


@require_tenant_role("COMPANY_ADMIN", "COMPANY_STAFF")
def admin_technician_list(request: HttpRequest, **kwargs) -> HttpResponse:
    """List all technicians for the company."""
    company = request.company
    technicians = Technician.objects.filter(company=company).select_related("user")
    return render(request, "tenants/admin_technicians.html", {
        "company": company, "technicians": technicians,
    })


@require_tenant_role("COMPANY_ADMIN")
def admin_technician_create(request: HttpRequest, **kwargs) -> HttpResponse:
    """Create a new technician (creates user + technician profile)."""
    company = request.company
    error = ""

    if request.method == "POST":
        form = TechnicianCreateForm(request.POST)
        if form.is_valid():
            from apps.common.phone_utils import normalize_iran_mobile

            username = form.cleaned_data["username"].strip().lower()
            phone = form.cleaned_data["phone"]
            normalized_phone = normalize_iran_mobile(phone) or phone

            # Validate username uniqueness
            if CompanyUser.objects.filter(username=username).exists():
                error = "این نام کاربری قبلاً استفاده شده است. لطفاً نام کاربری دیگری انتخاب کنید."
            elif not username:
                error = "نام کاربری الزامی است."
            else:
                _tech_password = form.cleaned_data.get("password") or "123456"
                user = CompanyUser.objects.create_user(
                    username=username,
                    password=_tech_password,
                    company=company,
                    role=UserRole.TECHNICIAN,
                    phone=normalized_phone,
                    first_name=form.cleaned_data["first_name"],
                    last_name=form.cleaned_data["last_name"],
                    must_change_password=(_tech_password == "123456"),
                )
                technician = Technician.objects.create(
                    company=company,
                    user=user,
                    is_available=form.cleaned_data.get("is_available", True),
                    service_wage_percent=_parse_wage_percent(request.POST.get("service_wage_percent", "0")),
                    goods_wage_percent=_parse_wage_percent(request.POST.get("goods_wage_percent", "0")),
                    travel_wage_percent=_parse_wage_percent(request.POST.get("travel_wage_percent", "0")),
                    shaba_number=_parse_shaba(request.POST.get("shaba_number", "")),
                )
                # Create legacy free-text skills
                skills_str = form.cleaned_data.get("skills", "")
                if skills_str:
                    for skill_name in skills_str.split(","):
                        skill_name = skill_name.strip()
                        if skill_name:
                            TechnicianSkill.objects.get_or_create(
                                company=company, technician=technician, name=skill_name,
                                defaults={"level": "intermediate"},
                            )
                _save_technician_category_priorities(
                    technician=technician, company=company, post_data=request.POST,
                )
                return redirect(f"/{company.code}/admin/technicians/")
    else:
        form = TechnicianCreateForm()

    from apps.tenants.models import CompanyServiceCategory
    categories = CompanyServiceCategory.objects.filter(company=company, is_active=True).order_by("sort_order", "title")
    category_priority_rows = [
        {"category": category, "priority": ""}
        for category in categories
    ]

    return render(request, "tenants/admin_technician_form.html", {
        "company": company, "form": form, "error": error, "is_edit": False,
        "category_priority_rows": category_priority_rows,
    })


@require_tenant_role("COMPANY_ADMIN")
def admin_technician_edit(request: HttpRequest, technician_id: int, **kwargs) -> HttpResponse:
    """Edit a technician."""
    company = request.company
    technician = Technician.objects.filter(id=technician_id, company=company).select_related("user").first()
    if not technician:
        raise Http404("تکنسین یافت نشد.")

    error = ""
    if request.method == "POST":
        form = TechnicianEditForm(request.POST)
        if form.is_valid():
            from apps.common.phone_utils import normalize_iran_mobile
            new_phone_raw = form.cleaned_data["phone"].strip()
            new_phone = normalize_iran_mobile(new_phone_raw) or new_phone_raw
            if CompanyUser.objects.filter(phone=new_phone).exclude(id=technician.user_id).exists():
                error = "کاربری دیگر با این شماره تلفن وجود دارد."
                return render(request, "tenants/admin_technician_form.html", {
                    "company": company, "form": form, "error": error, "is_edit": True, "technician": technician,
                    "category_priority_rows": [
                        {"category": category, "priority": request.POST.get(f"category_priority_{category.id}", "")}
                        for category in __import__("apps.tenants.models", fromlist=["CompanyServiceCategory"]).CompanyServiceCategory.objects.filter(company=company, is_active=True).order_by("sort_order", "title")
                    ],
                })

            technician.user.phone = new_phone
            technician.user.first_name = form.cleaned_data["first_name"]
            technician.user.last_name = form.cleaned_data["last_name"]
            new_password = form.cleaned_data.get("password", "").strip()
            if new_password:
                technician.user.set_password(new_password)
            technician.user.save()
            technician.is_available = form.cleaned_data.get("is_available", True)
            technician.service_wage_percent = _parse_wage_percent(request.POST.get("service_wage_percent", "0"))
            technician.goods_wage_percent = _parse_wage_percent(request.POST.get("goods_wage_percent", "0"))
            technician.travel_wage_percent = _parse_wage_percent(request.POST.get("travel_wage_percent", "0"))
            new_shaba = _parse_shaba(request.POST.get("shaba_number", ""))
            if new_shaba != technician.shaba_number:
                # SHABA changed — reset verification so platform owner re-reviews
                technician.shaba_number = new_shaba
                if new_shaba:
                    from apps.accounts.models import Technician as _Tech
                    technician.financial_verification_status = _Tech.FinancialVerificationStatus.PENDING
                    technician.shaba_verified = False
                    technician.shaba_verified_at = None
                else:
                    from apps.accounts.models import Technician as _Tech
                    technician.financial_verification_status = _Tech.FinancialVerificationStatus.NOT_SUBMITTED
                    technician.shaba_verified = False
                    technician.shaba_verified_at = None
            technician.save(update_fields=[
                "is_available", "service_wage_percent", "goods_wage_percent", "travel_wage_percent",
                "shaba_number", "shaba_verified", "shaba_verified_at", "financial_verification_status",
                "updated_at",
            ])
            # Update skills
            skills_str = form.cleaned_data.get("skills", "")
            TechnicianSkill.objects.filter(company=company, technician=technician).delete()
            if skills_str:
                for skill_name in skills_str.split(","):
                    skill_name = skill_name.strip()
                    if skill_name:
                        TechnicianSkill.objects.create(
                            company=company, technician=technician,
                            name=skill_name, level="intermediate",
                        )
            _save_technician_category_priorities(
                technician=technician, company=company, post_data=request.POST,
            )
            return redirect(f"/{company.code}/admin/technicians/")
    else:
        skills = ", ".join(
            TechnicianSkill.objects.filter(company=company, technician=technician).values_list("name", flat=True)
        )
        form = TechnicianEditForm(initial={
            "username": technician.user.username,
            "phone": technician.user.phone,
            "first_name": technician.user.first_name,
            "last_name": technician.user.last_name,
            "is_available": technician.is_available,
            "skills": skills,
        })

    from apps.tenants.models import CompanyServiceCategory
    categories = CompanyServiceCategory.objects.filter(company=company, is_active=True).order_by("sort_order", "title")
    existing_category_priorities = {
        row.category_id: row.priority
        for row in TechnicianCategorySkill.objects.filter(technician=technician)
    }
    category_priority_rows = [
        {"category": category, "priority": str(existing_category_priorities.get(category.id, ""))}
        for category in categories
    ]

    return render(request, "tenants/admin_technician_form.html", {
        "company": company, "form": form, "error": error, "is_edit": True, "technician": technician,
        "category_priority_rows": category_priority_rows,
    })


@require_tenant_role("COMPANY_ADMIN")
def admin_technician_delete(request: HttpRequest, technician_id: int, **kwargs) -> HttpResponse:
    """Deactivate a technician (soft delete by deactivating user)."""
    company = request.company
    technician = Technician.objects.filter(id=technician_id, company=company).first()
    if not technician:
        raise Http404("تکنسین یافت نشد.")

    if request.method == "POST":
        technician.is_available = False
        technician.save(update_fields=["is_available", "updated_at"])
        technician.user.is_active = False
        technician.user.save(update_fields=["is_active"])
        return redirect(f"/{company.code}/admin/technicians/")

    return render(request, "tenants/admin_technician_delete.html", {
        "company": company, "technician": technician,
    })


# =============================================================================
# CUSTOMERS LIST + DETAIL
# =============================================================================



@require_tenant_role("COMPANY_ADMIN", "COMPANY_STAFF")
def admin_customer_list(request: HttpRequest, **kwargs) -> HttpResponse:
    from django.db.models import Count, Max, Q

    company = request.company
    q = (request.GET.get("q") or "").strip()
    customers = Customer.objects.filter(company=company)

    if q:
        customers = customers.filter(
            Q(first_name__icontains=q) | Q(last_name__icontains=q) |
            Q(phone__icontains=q) | Q(email__icontains=q)
        )

    customers = customers.annotate(
        orders_count=Count("orders", filter=Q(orders__company=company), distinct=True),
        active_orders_count=Count("orders", filter=Q(orders__company=company) & ~Q(orders__status__in=[Order.Status.DONE, Order.Status.CANCELLED]), distinct=True),
        done_orders_count=Count("orders", filter=Q(orders__company=company, orders__status=Order.Status.DONE), distinct=True),
        cancelled_orders_count=Count("orders", filter=Q(orders__company=company, orders__status=Order.Status.CANCELLED), distinct=True),
        last_order_at=Max("orders__created_at", filter=Q(orders__company=company)),
    ).order_by("-last_order_at", "-created_at", "first_name", "last_name")

    customer_rows = []
    total_orders = active_orders = done_orders = 0
    for customer in customers:
        customer.full_name = _get_customer_display_name(customer) or customer.phone
        customer.addresses_count = Order.objects.filter(company=company, customer=customer).exclude(address__isnull=True).exclude(address="").values("address").distinct().count()
        total_orders += int(customer.orders_count or 0)
        active_orders += int(customer.active_orders_count or 0)
        done_orders += int(customer.done_orders_count or 0)
        customer_rows.append(customer)

    stats = {
        "total_customers": len(customer_rows),
        "total_orders": total_orders,
        "active_orders": active_orders,
        "done_orders": done_orders,
    }

    return render(request, "tenants/admin_customers.html", {
        "company": company, "customers": customer_rows, "q": q, "stats": stats,
    })


def _build_customer_invoice_query(company, customer: Customer):
    from django.db.models import Q

    field_names = {field.name for field in Invoice._meta.fields}
    invoice_q = Q()
    has_condition = False

    if "customer" in field_names:
        invoice_q |= Q(customer=customer)
        has_condition = True

    phone = getattr(customer, "phone", "") or ""
    for phone_field in ("customer_phone", "display_customer_phone", "customer_phone_snapshot", "phone_number"):
        if phone and phone_field in field_names:
            invoice_q |= Q(**{phone_field: phone})
            has_condition = True

    if not has_condition:
        return Invoice.objects.none()

    return Invoice.objects.filter(company=company).filter(invoice_q).distinct()


@require_tenant_role("COMPANY_ADMIN", "COMPANY_STAFF")
def admin_customer_detail(request: HttpRequest, customer_id: int, **kwargs) -> HttpResponse:
    from django.db.models import Q

    company = request.company
    customer = Customer.objects.filter(id=customer_id, company=company).first()
    if not customer:
        raise Http404("مشتری یافت نشد.")

    customer_display_name = _get_customer_display_name(customer) or customer.phone
    phone = customer.phone or ""

    orders = (
        Order.objects.filter(company=company)
        .filter(Q(customer=customer) | Q(customer_phone=phone))
        .select_related("technician", "technician__user", "service_category")
        .order_by("-created_at")
        .distinct()
    )
    invoices = _build_customer_invoice_query(company=company, customer=customer).order_by("-created_at")
    service_requests = ServiceRequest.objects.filter(company=company, customer_phone=phone).select_related("service", "order").order_by("-created_at")

    seen_addresses = set()
    addresses = []

    def add_address(address, last_order_id=None, last_used_at=None):
        address = (address or "").strip()
        if not address or address in seen_addresses:
            return
        seen_addresses.add(address)
        addresses.append({"address": address, "last_order_id": last_order_id, "last_used_at": last_used_at})

    add_address(getattr(customer, "address", ""), None, getattr(customer, "updated_at", None))
    for order in orders:
        add_address(order.address, order.id, order.created_at)
    for req in service_requests:
        add_address(req.address, req.order_id, req.created_at)

    sms_messages = []
    try:
        from apps.sms.models import SMSOutbox
        sms_messages = list(SMSOutbox.objects.filter(company=company, phone_number=phone).order_by("-created_at")[:50])
    except Exception:
        sms_messages = []

    stats = {
        "orders_count": orders.count(),
        "done_orders_count": orders.filter(status=Order.Status.DONE).count(),
        "cancelled_orders_count": orders.filter(status=Order.Status.CANCELLED).count(),
        "invoices_count": invoices.count(),
        "service_requests_count": service_requests.count(),
        "sms_count": len(sms_messages),
    }

    return render(request, "tenants/admin_customer_detail.html", {
        "company": company,
        "customer": customer,
        "customer_display_name": customer_display_name,
        "orders": orders[:100],
        "invoices": invoices[:100],
        "service_requests": service_requests[:50],
        "sms_messages": sms_messages,
        "addresses": addresses,
        "stats": stats,
    })


# =============================================================================
# ORDERS ADMIN
# =============================================================================



@require_tenant_role("COMPANY_ADMIN", "COMPANY_STAFF")
def admin_order_list(request: HttpRequest, **kwargs) -> HttpResponse:
    """List all orders with status filtering."""
    company = request.company
    status_filter = request.GET.get("status", "")

    if status_filter:
        orders = OrderSelector.get_by_status(company=company, status=status_filter)
    else:
        orders = OrderSelector.get_for_company(company=company)

    return render(request, "tenants/admin_orders.html", {
        "company": company, "orders": orders, "status_filter": status_filter,
        "statuses": Order.Status.choices,
    })


@require_tenant_role("COMPANY_ADMIN", "COMPANY_STAFF")
def admin_order_detail(request: HttpRequest, order_id: int, **kwargs) -> HttpResponse:
    """Order detail with admin actions (assign, cancel)."""
    from apps.orders.item_services import OrderItemService

    company = request.company
    order = OrderSelector.get_by_id_for_company(order_id=order_id, company=company)
    if not order:
        raise Http404("سفارش یافت نشد.")


    prefill_customer_phone = (request.GET.get("customer_phone") or request.GET.get("phone") or "").strip()
    prefill_customer_name = ""
    prefill_customer_address = ""

    customer_id_raw = request.GET.get("customer_id")
    prefill_customer = None
    if customer_id_raw:
        try:
            prefill_customer = Customer.objects.filter(company=company, id=int(customer_id_raw)).first()
        except (TypeError, ValueError):
            prefill_customer = None

    if prefill_customer is None and prefill_customer_phone:
        try:
            from apps.common.phone_utils import normalize_iran_mobile
            normalized_phone = normalize_iran_mobile(prefill_customer_phone)
            if normalized_phone:
                prefill_customer = Customer.objects.filter(company=company, phone=normalized_phone).first()
                prefill_customer_phone = normalized_phone
        except Exception:
            pass

    if prefill_customer is not None:
        prefill_customer_phone = prefill_customer.phone
        prefill_customer_name = _get_customer_display_name(prefill_customer)
        prefill_customer_address = getattr(prefill_customer, "address", "") or ""
        last_order_with_address = Order.objects.filter(company=company, customer=prefill_customer).exclude(address__isnull=True).exclude(address="").order_by("-created_at").first()
        if last_order_with_address:
            prefill_customer_address = last_order_with_address.address

    technicians = Technician.objects.filter(company=company, is_available=True)
    error = ""

    # Handle assign technician
    if request.method == "POST" and "assign_technician" in request.POST:
        tech_id = request.POST.get("technician_id")
        if tech_id:
            from apps.orders.services import OrderAssignService
            technician = Technician.objects.filter(id=tech_id, company=company).first()
            if technician:
                try:
                    OrderAssignService.assign(
                        order=order, technician=technician, assigned_by=request.user,
                    )
                    return redirect(f"/{company.code}/admin/orders/{order.id}/")
                except ValueError as e:
                    error = str(e)

    # Handle force cancel
    if request.method == "POST" and "force_cancel" in request.POST:
        reason = request.POST.get("cancel_reason", "")
        try:
            OrderCancelService.force_cancel(
                order=order, cancelled_by=request.user, reason=reason,
            )
            return redirect(f"/{company.code}/admin/orders/{order.id}/")
        except ValueError as e:
            error = str(e)

    # Handle confirm cancel (CANCEL_REQUESTED → CANCELLED)
    if request.method == "POST" and "confirm_cancel" in request.POST:
        if order.status == Order.Status.CANCEL_REQUESTED:
            try:
                OrderCancelService.force_cancel(
                    order=order, cancelled_by=request.user,
                    reason="Admin confirmed cancellation request.",
                )
                return redirect(f"/{company.code}/admin/orders/{order.id}/")
            except ValueError as e:
                error = str(e)

    # Handle recycle/reopen (cancel old + create replacement NEW)
    if request.method == "POST" and "recycle_order" in request.POST:
        from apps.orders.recycle_service import OrderRecycleService
        if order.status in [Order.Status.CANCEL_REQUESTED, Order.Status.CANCELLED]:
            # Already cancelled — just recycle
            pass
        try:
            new_order = OrderRecycleService.recycle(
                order=order, recycled_by=request.user,
            )
            return redirect(f"/{company.code}/admin/orders/{new_order.id}/")
        except ValueError as e:
            error = str(e)

    # Permission flags for template action visibility
    from apps.accounts.operator_access import is_company_admin as _is_admin, operator_has_permission
    user = request.user
    user_is_admin = _is_admin(user)
    can_edit_order = user_is_admin or operator_has_permission(company=company, operator=user, permission_key="admin_order_edit")
    can_assign_order = user_is_admin or operator_has_permission(company=company, operator=user, permission_key="admin_order_assign")
    can_force_cancel = user_is_admin or operator_has_permission(company=company, operator=user, permission_key="admin_cancel_request_approve")
    can_create_invoice = user_is_admin or operator_has_permission(company=company, operator=user, permission_key="admin_invoice_create_from_order")
    can_return_to_cycle = user_is_admin or operator_has_permission(company=company, operator=user, permission_key="admin_order_return_to_cycle")

    # Check if active (non-cancelled) invoice already exists for this order
    has_active_invoice = Invoice.objects.filter(company=company, order=order).exclude(status=Invoice.Status.CANCELLED).exists()

    return render(request, "tenants/admin_order_detail.html", {
        "company": company, "order": order, "technicians": technicians, "error": error,
        "item_values": OrderItemService.get_values_display(order=order),
        "service_date_jalali": format_jalali_date(order.service_date),
        "can_edit_order": can_edit_order,
        "can_assign_order": can_assign_order,
        "can_force_cancel": can_force_cancel,
        "can_create_invoice": can_create_invoice,
        "can_return_to_cycle": can_return_to_cycle,
        "has_active_invoice": has_active_invoice,
    })



def _parse_invoice_items_from_post(request: HttpRequest) -> list[dict]:
    """Parse invoice item rows from admin invoice form."""
    descriptions = request.POST.getlist("item_description")
    quantities = request.POST.getlist("item_quantity")
    unit_prices = request.POST.getlist("item_unit_price")
    discounts = request.POST.getlist("item_discount_amount")

    items = []
    max_len = max(len(descriptions), len(quantities), len(unit_prices), len(discounts), 0)
    for i in range(max_len):
        description = descriptions[i] if i < len(descriptions) else ""
        if not (description or "").strip():
            continue
        items.append({
            "description": description,
            "quantity": quantities[i] if i < len(quantities) else "1",
            "unit_price": unit_prices[i] if i < len(unit_prices) else "0",
            "discount_amount": discounts[i] if i < len(discounts) else "0",
        })
    return items


# =============================================================================
# INVOICES ADMIN
# =============================================================================


@require_tenant_role("COMPANY_ADMIN", "COMPANY_STAFF")
def admin_invoice_list(request: HttpRequest, **kwargs) -> HttpResponse:
    """List all invoices with status filtering."""
    company = request.company
    status_filter = request.GET.get("status", "")

    if status_filter:
        invoices = InvoiceSelector.get_by_status(company=company, status=status_filter)
    else:
        invoices = InvoiceSelector.get_for_company(company=company)

    return render(request, "tenants/admin_invoices.html", {
        "company": company, "invoices": invoices, "status_filter": status_filter,
        "statuses": Invoice.Status.choices,
    })


@require_tenant_role("COMPANY_ADMIN", "COMPANY_STAFF")
def admin_invoice_create_from_order(request: HttpRequest, order_id: int, **kwargs) -> HttpResponse:
    """Create a draft invoice from an order and redirect to edit page."""
    company = request.company
    order = OrderSelector.get_by_id_for_company(order_id=order_id, company=company)
    if not order:
        raise Http404("سفارش یافت نشد.")

    invoice = InvoiceCreateService.create_from_order(order=order, created_by=request.user)
    return redirect(f"/{company.code}/admin/invoices/{invoice.id}/edit/")


@require_tenant_role("COMPANY_ADMIN", "COMPANY_STAFF")
def admin_invoice_detail(request: HttpRequest, invoice_id: int, **kwargs) -> HttpResponse:
    """Invoice detail with admin actions (issue, cancel, mark paid)."""
    company = request.company
    invoice = InvoiceSelector.get_by_id_for_company(invoice_id=invoice_id, company=company)
    if not invoice:
        raise Http404("فاکتور یافت نشد.")

    error = ""

    if request.method == "POST" and "issue_invoice" in request.POST:
        try:
            InvoiceIssueService.issue(invoice=invoice)
            return redirect(f"/{company.code}/admin/invoices/{invoice.id}/")
        except ValueError as e:
            error = str(e)

    if request.method == "POST" and "mark_paid_company_cash" in request.POST:
        try:
            if not invoice.is_payable:
                raise ValueError("این فاکتور قابل پرداخت نیست. فقط فاکتورهای صادرشده قابل ثبت پرداخت هستند.")
            from apps.payments.models import Payment
            payment = Payment.objects.create(
                company=company,
                invoice=invoice,
                amount=invoice.total_amount,
                status=Payment.Status.PAID,
                metadata={"payment_source": "CASH_RECEIVED_BY_COMPANY", "method": "cash"},
            )
            InvoiceMarkPaidService.mark_paid(invoice=invoice, payment=payment, payment_method="cash")
            return redirect(f"/{company.code}/admin/invoices/{invoice.id}/")
        except (ValueError, Exception) as e:
            error = str(e)

    if request.method == "POST" and "mark_paid_technician_cash" in request.POST:
        try:
            if not invoice.is_payable:
                raise ValueError("این فاکتور قابل پرداخت نیست. فقط فاکتورهای صادرشده قابل ثبت پرداخت هستند.")
            from apps.payments.models import Payment
            order = getattr(invoice, "order", None)
            technician_id = getattr(getattr(order, "technician", None), "id", None)
            payment = Payment.objects.create(
                company=company,
                invoice=invoice,
                amount=invoice.total_amount,
                status=Payment.Status.PAID,
                metadata={
                    "payment_source": "CASH_RECEIVED_BY_TECHNICIAN",
                    "method": "cash",
                    "technician_id": technician_id,
                },
            )
            InvoiceMarkPaidService.mark_paid(invoice=invoice, payment=payment, payment_method="cash")
            return redirect(f"/{company.code}/admin/invoices/{invoice.id}/")
        except (ValueError, Exception) as e:
            error = str(e)

    if request.method == "POST" and "cancel_invoice" in request.POST:
        reason = request.POST.get("cancel_reason", "")
        try:
            InvoiceCancelService.cancel(invoice=invoice, reason=reason)
            return redirect(f"/{company.code}/admin/invoices/{invoice.id}/")
        except ValueError as e:
            error = str(e)

    public_url = request.build_absolute_uri(
        f"/{company.code}/invoices/public/{invoice.public_code}/"
    ) if invoice.public_code else ""

    # Wage calculation for display
    from apps.invoices.services_wage import calculate_technician_wage
    from apps.payments.selectors import PaymentSelector

    wage_breakdown = calculate_technician_wage(invoice)
    paid_payment = PaymentSelector.get_latest_paid_for_invoice(company=company, invoice_id=invoice.id)
    payment_info = PaymentSelector.build_display_info(paid_payment)

    from apps.invoices.services_preview import InvoiceFinancialPreviewService
    financial_preview = InvoiceFinancialPreviewService.compute(invoice)

    return render(request, "tenants/admin_invoice_detail.html", {
        "company": company, "invoice": invoice, "error": error,
        "public_url": public_url,
        "wage": wage_breakdown,
        "paid_payment": paid_payment,
        "payment_info": payment_info,
        "financial_preview": financial_preview,
    })


@require_tenant_role("COMPANY_ADMIN", "COMPANY_STAFF")
def admin_invoice_edit(request: HttpRequest, invoice_id: int, **kwargs) -> HttpResponse:
    """Edit a draft invoice header and rows."""
    company = request.company
    invoice = InvoiceSelector.get_by_id_for_company(invoice_id=invoice_id, company=company)
    if not invoice:
        raise Http404("فاکتور یافت نشد.")

    error = ""
    if request.method == "POST":
        try:
            InvoiceUpdateService.update(
                invoice=invoice,
                data=request.POST,
                items=_parse_invoice_items_from_post(request),
            )
            return redirect(f"/{company.code}/admin/invoices/{invoice.id}/")
        except ValueError as e:
            error = str(e)

    existing_items = list(invoice.items.all())
    blank_rows_count = max(5, 10 - len(existing_items))

    return render(request, "tenants/admin_invoice_edit.html", {
        "company": company,
        "invoice": invoice,
        "items": existing_items,
        "blank_rows": range(blank_rows_count),
        "error": error,
    })


@require_tenant_role("COMPANY_ADMIN", "COMPANY_STAFF")
def admin_invoice_print(request: HttpRequest, public_code: str, **kwargs) -> HttpResponse:
    """Print/download invoice as clean HTML page (Phase 27A). Uses public_code URL."""
    company = request.company
    invoice = InvoiceSelector.get_by_public_code_for_company(company=company, public_code=public_code)
    if not invoice:
        raise Http404("فاکتور یافت نشد.")
    if invoice.status == Invoice.Status.CANCELLED:
        raise Http404("فاکتور قابل چاپ نیست.")
    return render(request, "invoices/print.html", {
        "invoice": invoice,
        "company": company,
    })


# =============================================================================
# REQUESTS (existing)
# =============================================================================


@require_tenant_role("COMPANY_ADMIN", "COMPANY_STAFF")
def admin_request_list(request: HttpRequest, **kwargs) -> HttpResponse:
    """Admin view for service requests."""
    company = request.company
    requests = ServiceRequestSelector.get_for_company(company=company)
    return render(request, "tenants/admin_requests.html", {
        "company": company, "requests": requests,
    })



# =============================================================================
# ORDER EDIT + ASSIGN + CREATE
# =============================================================================


def _parse_money_from_post(request: HttpRequest, field_name: str) -> int:
    raw = normalize_digits(request.POST.get(field_name) or "0").replace(",", "").strip()
    if raw == "":
        return 0
    try:
        return int(float(raw))
    except ValueError as exc:
        raise ValueError(f"مقدار {field_name} باید عدد باشد.") from exc


def _parse_wage_percent(value, default=0):
    """Parse wage percentage from form input (0-100, allows decimals)."""
    from decimal import Decimal, InvalidOperation
    if not value or str(value).strip() == "":
        return Decimal(str(default))
    try:
        result = Decimal(str(value).replace(",", "").strip())
        if result < 0:
            return Decimal("0")
        if result > 100:
            return Decimal("100")
        return result
    except (InvalidOperation, ValueError, TypeError):
        return Decimal(str(default))


def _parse_shaba(value: str) -> str:
    """Normalize and validate SHABA format: IR + 24 digits, max 26 chars."""
    v = (value or "").strip().upper().replace(" ", "").replace("-", "")
    if not v:
        return ""
    if len(v) > 26:
        v = v[:26]
    return v


def _get_customer_display_name(customer: Customer | None) -> str:
    if customer is None:
        return ""
    return f"{customer.first_name} {customer.last_name}".strip()


@require_tenant_role("COMPANY_ADMIN", "COMPANY_STAFF")
def admin_customer_lookup(request: HttpRequest, **kwargs) -> HttpResponse:
    """
    JSON endpoint for customer phone lookup during order creation.

    GET /<company_code>/admin/customers/lookup/?phone=09121234567

    Returns JSON with customer info and previous addresses if found.
    Company-scoped: never leaks cross-tenant data.
    """
    from django.http import JsonResponse
    from apps.common.phone_utils import normalize_iran_mobile

    company = request.company
    phone_raw = request.GET.get("phone", "").strip()
    phone = normalize_iran_mobile(phone_raw)

    if not phone:
        return JsonResponse({
            "ok": False,
            "exists": False,
            "error": "شماره موبایل معتبر نیست" if phone_raw else "",
        })

    customer = Customer.objects.filter(company=company, phone=phone).first()

    if customer is None:
        return JsonResponse({
            "ok": True,
            "exists": False,
            "phone": phone,
            "addresses": [],
        })

    # Get previous addresses from this customer's orders (deduplicated, most recent first)
    from apps.orders.models import Order
    addresses = (
        Order.objects
        .filter(company=company, customer=customer)
        .exclude(address__isnull=True)
        .exclude(address="")
        .values_list("address", "id", "created_at")
        .order_by("-created_at")
    )

    seen = set()
    address_list = []
    for addr, order_id, created_at in addresses:
        addr_clean = addr.strip()
        if addr_clean and addr_clean not in seen:
            seen.add(addr_clean)
            address_list.append({
                "address": addr_clean,
                "last_order_id": order_id,
                "last_used_at": created_at.isoformat() if created_at else "",
            })
            if len(address_list) >= 10:
                break

    return JsonResponse({
        "ok": True,
        "exists": True,
        "phone": phone,
        "customer": {
            "id": customer.id,
            "first_name": customer.first_name,
            "last_name": customer.last_name,
            "full_name": f"{customer.first_name} {customer.last_name}".strip(),
        },
        "addresses": address_list,
    })


@require_tenant_role("COMPANY_ADMIN", "COMPANY_STAFF")
def admin_order_create(request: HttpRequest, **kwargs) -> HttpResponse:
    """Create a new order from admin panel."""
    import json
    from apps.orders.services import OrderCreateByAdminService
    from apps.orders.item_services import OrderItemService
    from apps.tenants.selectors import CompanyServiceCategorySelector

    company = request.company
    error = ""

    if request.method == "POST":
        try:
            from apps.common.phone_utils import normalize_iran_mobile

            customer_name = request.POST.get("customer_name", "").strip()
            customer_phone_raw = request.POST.get("customer_phone", "").strip()
            customer_phone = normalize_iran_mobile(customer_phone_raw)

            if not customer_phone and customer_phone_raw:
                raise ValueError("شماره موبایل معتبر نیست. فرمت صحیح: 09xxxxxxxxx")
            if not customer_phone:
                raise ValueError("شماره تماس مشتری الزامی است.")

            service_date_raw = request.POST.get("service_date_jalali", "").strip()
            service_date = parse_jalali_date(service_date_raw)
            if not service_date and service_date_raw:
                raise ValueError("تاریخ وارد شده معتبر نیست. فرمت صحیح: 1404/03/15")

            # Customer find-or-create with normalized phone
            customer = Customer.objects.filter(company=company, phone=customer_phone).first()
            if customer is None:
                # Create new customer
                parts = customer_name.split(" ", 1) if customer_name else ["", ""]
                customer = Customer.objects.create(
                    company=company,
                    first_name=parts[0],
                    last_name=parts[1] if len(parts) > 1 else "",
                    phone=customer_phone,
                    address=request.POST.get("address", ""),
                )
            elif customer_name:
                # Update customer name if admin corrected it
                parts = customer_name.split(" ", 1)
                new_first = parts[0]
                new_last = parts[1] if len(parts) > 1 else ""
                if new_first and (new_first != customer.first_name or new_last != customer.last_name):
                    customer.first_name = new_first
                    customer.last_name = new_last
                    customer.save(update_fields=["first_name", "last_name", "updated_at"])

            order = OrderCreateByAdminService.create(
                company=company,
                created_by=request.user,
                title=request.POST.get("title", ""),
                customer=customer,
                customer_name=customer_name or f"{customer.first_name} {customer.last_name}".strip(),
                customer_phone=customer_phone,
                description=request.POST.get("description", ""),
                address=request.POST.get("address", ""),
                service_date=service_date,
                priority="normal",
                price_estimate=0,
                extra_payment=_parse_money_from_post(request, "extra_payment"),
                wage_deduction=_parse_money_from_post(request, "wage_deduction"),
                required_skill="",
                internal_note=request.POST.get("internal_note", ""),
                status=request.POST.get("status", Order.Status.NEW),
                service_category_id=int(request.POST.get("service_category_id") or 0) or None,
                service_subcategory_id=None,
                technician_id=int(request.POST.get("technician_id") or 0) or None,
            )
            # Save dynamic item values
            OrderItemService.save_items_from_post(
                order=order, post_data=request.POST, company=company,
            )
            return redirect(f"/{company.code}/admin/orders/")
        except ValueError as e:
            error = str(e)

    technicians = Technician.objects.filter(company=company, is_available=True)
    categories = CompanyServiceCategorySelector.get_active_for_company(company=company)
    item_definitions_json = json.dumps(
        OrderItemService.get_definitions_json(company=company)
    )

    # PATCH 7C9: ensure order-create customer prefill variables always exist.
    # Some previous patches added these variables to the template context, but on
    # normal GET /admin/orders/create/ they may not have been initialized.
    if (
        "prefill_customer_phone" not in locals()
        or "prefill_customer_name" not in locals()
        or "prefill_customer_address" not in locals()
    ):
        prefill_customer_phone = (request.GET.get("customer_phone") or request.GET.get("phone") or "").strip()
        prefill_customer_name = ""
        prefill_customer_address = ""

        customer_id_raw = request.GET.get("customer_id")
        prefill_customer = None

        if customer_id_raw:
            try:
                from apps.accounts.models import Customer as PrefillCustomer

                prefill_customer = PrefillCustomer.objects.filter(
                    company=company,
                    id=int(customer_id_raw),
                ).first()
            except (TypeError, ValueError):
                prefill_customer = None

        if prefill_customer is None and prefill_customer_phone:
            try:
                from apps.common.phone_utils import normalize_iran_mobile
                from apps.accounts.models import Customer as PrefillCustomer

                normalized_phone = normalize_iran_mobile(prefill_customer_phone)
                if normalized_phone:
                    prefill_customer_phone = normalized_phone
                    prefill_customer = PrefillCustomer.objects.filter(
                        company=company,
                        phone=normalized_phone,
                    ).first()
            except Exception:
                pass

        if prefill_customer is not None:
            prefill_customer_phone = getattr(prefill_customer, "phone", "") or prefill_customer_phone

            try:
                prefill_customer_name = _get_customer_display_name(prefill_customer)
            except Exception:
                prefill_customer_name = (
                    f"{getattr(prefill_customer, 'first_name', '')} {getattr(prefill_customer, 'last_name', '')}".strip()
                    or getattr(prefill_customer, "phone", "")
                )

            prefill_customer_address = getattr(prefill_customer, "address", "") or ""

            try:
                from apps.orders.models import Order as PrefillOrder

                last_order_with_address = (
                    PrefillOrder.objects
                    .filter(company=company, customer=prefill_customer)
                    .exclude(address__isnull=True)
                    .exclude(address="")
                    .order_by("-created_at")
                    .first()
                )
                if last_order_with_address:
                    prefill_customer_address = last_order_with_address.address
            except Exception:
                pass



    return render(request, "tenants/admin_order_create.html", {
        "company": company, "technicians": technicians,
        "categories": categories,
        "item_definitions_json": item_definitions_json,
        "today_jalali": today_jalali_date(),
        "prefill_customer_phone": prefill_customer_phone,
        "prefill_customer_name": prefill_customer_name,
        "prefill_customer_address": prefill_customer_address,
        "error": error, "statuses": Order.Status.choices,
    })


@require_tenant_role("COMPANY_ADMIN", "COMPANY_STAFF")
def admin_order_edit(request: HttpRequest, order_id: int, **kwargs) -> HttpResponse:
    """Edit order details (admin/staff only), including category/subcategory."""
    import json
    from apps.orders.services import OrderUpdateService
    from apps.orders.item_services import OrderItemService
    from apps.tenants.selectors import CompanyServiceCategorySelector

    company = request.company
    order = OrderSelector.get_by_id_for_company(order_id=order_id, company=company)
    if not order:
        raise Http404("سفارش یافت نشد.")

    error = ""
    if request.method == "POST":
        try:
            from apps.common.phone_utils import normalize_iran_mobile

            customer_phone_raw = request.POST.get("customer_phone", "").strip()
            customer_phone = normalize_iran_mobile(customer_phone_raw)
            if not customer_phone and customer_phone_raw:
                raise ValueError("شماره موبایل معتبر نیست. فرمت صحیح: 09xxxxxxxxx")

            data = {
                "title": request.POST.get("title", order.title),
                "customer_name": request.POST.get("customer_name", "").strip(),
                "customer_phone": customer_phone or request.POST.get("customer_phone", "").strip(),
                "description": request.POST.get("description", ""),
                "address": request.POST.get("address", ""),
                "service_date": parse_jalali_date(
                    request.POST.get("service_date_jalali", ""),
                ),
                "extra_payment": _parse_money_from_post(request, "extra_payment"),
                "wage_deduction": _parse_money_from_post(request, "wage_deduction"),
                "required_skill": "",
                "price_estimate": 0,
                "final_price": order.final_price,
                "internal_note": request.POST.get("internal_note", ""),
                "scheduled_for": None,
                "status": request.POST.get("status", order.status),
                "service_category_id": int(request.POST.get("service_category_id") or 0) or None,
                "service_subcategory_id": None,
            }
            customer_phone_norm = data["customer_phone"]
            if not data["customer_name"]:
                raise ValueError("نام مشتری الزامی است.")
            if not customer_phone_norm:
                raise ValueError("شماره تماس مشتری الزامی است.")
            customer = Customer.objects.filter(company=company, phone=customer_phone_norm).first()
            if customer is None:
                parts = data["customer_name"].split(" ", 1)
                customer = Customer.objects.create(
                    company=company,
                    first_name=parts[0],
                    last_name=parts[1] if len(parts) > 1 else "",
                    phone=customer_phone_norm,
                    address=data["address"],
                )
            data["customer_id"] = customer.id

            OrderUpdateService.update(order=order, updated_by=request.user, data=data)
            order.refresh_from_db()
            OrderItemService.save_items_from_post(
                order=order, post_data=request.POST, company=company,
            )
            technician_id = int(request.POST.get("technician_id") or 0) or None
            from apps.orders.services import OrderEditAssignService
            OrderEditAssignService.handle_assignment(
                order=order,
                technician_id=technician_id,
                assigned_by=request.user,
                company=company,
            )
            return redirect(f"/{company.code}/admin/orders/{order.id}/")
        except ValueError as e:
            error = str(e)


    technicians = Technician.objects.filter(company=company, is_available=True)
    categories = CompanyServiceCategorySelector.get_active_for_company(company=company)
    item_definitions_json = json.dumps(
        OrderItemService.get_definitions_json(company=company)
    )
    # Load existing item values for pre-populating the form
    existing_values = OrderItemService.get_existing_values(order=order)
    # Build a JSON-friendly dict: {definition_id: raw_value}
    existing_item_values = {}
    for def_id, val_obj in existing_values.items():
        if val_obj.item.kind in ("number", "money"):
            existing_item_values[def_id] = str(val_obj.value_number) if val_obj.value_number is not None else ""
        elif val_obj.item.kind == "text":
            existing_item_values[def_id] = val_obj.value_text or ""
        elif val_obj.item.kind == "bool":
            existing_item_values[def_id] = val_obj.value_bool if val_obj.value_bool is not None else False
    existing_item_values_json = json.dumps(existing_item_values)

    return render(request, "tenants/admin_order_edit.html", {
        "company": company, "order": order,
        "technicians": technicians, "categories": categories,
        "item_definitions_json": item_definitions_json,
        "existing_item_values_json": existing_item_values_json,
        "service_date_jalali": format_jalali_date(order.service_date),
        "statuses": Order.Status.choices,
        "error": error,
    })


@require_tenant_role("COMPANY_ADMIN", "COMPANY_STAFF")
def admin_order_assign(request: HttpRequest, order_id: int, **kwargs) -> HttpResponse:
    """Assign a technician to an order (admin manual assignment)."""
    from apps.orders.services import OrderAssignService

    company = request.company
    order = OrderSelector.get_by_id_for_company(order_id=order_id, company=company)
    if not order:
        raise Http404("سفارش یافت نشد.")

    error = ""
    if request.method == "POST":
        tech_id = request.POST.get("technician_id")
        if tech_id:
            technician = Technician.objects.filter(id=tech_id, company=company).first()
            if technician:
                try:
                    OrderAssignService.assign(
                        order=order, technician=technician, assigned_by=request.user,
                    )
                    return redirect(f"/{company.code}/admin/orders/{order.id}/")
                except ValueError as e:
                    error = str(e)
            else:
                error = "تکنسین یافت نشد."
        else:
            error = "لطفا یک تکنسین انتخاب کنید."

    technicians = Technician.objects.filter(company=company, is_available=True)
    return render(request, "tenants/admin_order_assign.html", {
        "company": company, "order": order, "technicians": technicians, "error": error,
    })



# =============================================================================
# BASE DATA MANAGEMENT
# =============================================================================


def _parse_positive_int(value, default=0):
    """Parse an integer safely for sort/priority fields."""
    try:
        if value in (None, ""):
            return default
        return max(0, int(value))
    except (TypeError, ValueError):
        return default


def _operator_role_choices():
    return [
        (UserRole.COMPANY_STAFF, "اپراتور"),
        (UserRole.COMPANY_ADMIN, "مدیر شرکت"),
    ]


@require_tenant_role("COMPANY_ADMIN", "COMPANY_STAFF")
def admin_base_data_home(request: HttpRequest, **kwargs) -> HttpResponse:
    """Base-data landing page for tenant admins."""
    company = request.company
    from apps.orders.models import OrderItemDefinition
    from apps.tenants.models import CompanyServiceCategory

    counts = {
        "categories": CompanyServiceCategory.objects.filter(company=company).count(),
        "items": OrderItemDefinition.objects.filter(company=company).count(),
        "technicians": Technician.objects.filter(company=company).count(),
        "operators": CompanyUser.objects.filter(
            company=company,
            role__in=[UserRole.COMPANY_ADMIN, UserRole.COMPANY_STAFF],
        ).count(),
    }
    return render(request, "tenants/admin_base_data.html", {
        "company": company,
        "counts": counts,
    })


# -----------------------------------------------------------------------------
# Service categories
# -----------------------------------------------------------------------------


@require_tenant_role("COMPANY_ADMIN")
def admin_base_categories(request: HttpRequest, **kwargs) -> HttpResponse:
    # List service categories with their order item definitions on the same page.
    company = request.company
    from apps.orders.models import OrderItemDefinition
    from apps.tenants.models import CompanyServiceCategory

    categories = CompanyServiceCategory.objects.filter(company=company).order_by("sort_order", "title")
    items = OrderItemDefinition.objects.filter(company=company).select_related("category").order_by(
        "category__sort_order", "category__title", "sort_order", "title",
    )

    items_by_category: dict[int, list] = {}
    for item in items:
        items_by_category.setdefault(item.category_id, []).append(item)

    category_rows = []
    for category in categories:
        category_rows.append({
            "category": category,
            "items": items_by_category.get(category.id, []),
        })

    return render(request, "tenants/admin_base_categories.html", {
        "company": company,
        "category_rows": category_rows,
        "categories": categories,
    })


@require_tenant_role("COMPANY_ADMIN")
def admin_base_category_create(request: HttpRequest, **kwargs) -> HttpResponse:
    company = request.company
    from apps.tenants.models import CompanyServiceCategory

    error = ""
    if request.method == "POST":
        title = request.POST.get("title", "").strip()
        if not title:
            error = "عنوان رسته الزامی است."
        else:
            category = CompanyServiceCategory.objects.create(
                company=company,
                title=title,
                description=request.POST.get("description", "").strip(),
                sort_order=_parse_positive_int(request.POST.get("sort_order"), 0),
                is_active=bool(request.POST.get("is_active")),
            )
            return redirect(f"/{company.code}/admin/base-data/categories/")

    return render(request, "tenants/admin_base_category_form.html", {
        "company": company,
        "category": None,
        "is_edit": False,
        "error": error,
    })


@require_tenant_role("COMPANY_ADMIN")
def admin_base_category_edit(request: HttpRequest, category_id: int, **kwargs) -> HttpResponse:
    company = request.company
    from apps.tenants.models import CompanyServiceCategory
    from apps.orders.models import OrderItemDefinition

    category = CompanyServiceCategory.objects.filter(id=category_id, company=company).first()
    if category is None:
        raise Http404("رسته خدمات یافت نشد.")

    error = ""
    if request.method == "POST":
        title = request.POST.get("title", "").strip()
        if not title:
            error = "عنوان رسته الزامی است."
        else:
            category.title = title
            category.description = request.POST.get("description", "").strip()
            category.sort_order = _parse_positive_int(request.POST.get("sort_order"), category.sort_order)
            category.is_active = bool(request.POST.get("is_active"))
            category.save(update_fields=["title", "description", "sort_order", "is_active", "updated_at"])
            return redirect(f"/{company.code}/admin/base-data/categories/")

    items = OrderItemDefinition.objects.filter(
        company=company, category=category,
    ).order_by("sort_order", "title")
    return render(request, "tenants/admin_base_category_form.html", {
        "company": company,
        "category": category,
        "items": items,
        "is_edit": True,
        "error": error,
    })



@require_tenant_role("COMPANY_ADMIN")
def admin_base_category_toggle_active(request: HttpRequest, category_id: int, **kwargs) -> HttpResponse:
    # Activate/deactivate a service category without deleting it.
    company = request.company
    from apps.tenants.models import CompanyServiceCategory

    category = CompanyServiceCategory.objects.filter(id=category_id, company=company).first()
    if category is None:
        raise Http404("رسته خدمات یافت نشد.")

    if request.method == "POST":
        action = request.POST.get("action", "")
        if action == "activate":
            category.is_active = True
        elif action == "deactivate":
            category.is_active = False
        else:
            category.is_active = not category.is_active
        category.save(update_fields=["is_active", "updated_at"])

    return redirect(f"/{company.code}/admin/base-data/categories/")


@require_tenant_role("COMPANY_ADMIN")
def admin_base_category_delete(request: HttpRequest, category_id: int, **kwargs) -> HttpResponse:
    # Hard-delete a service category. Deactivate is handled separately.
    company = request.company
    from apps.orders.models import Order, OrderItemDefinition
    from apps.tenants.models import CompanyServiceCategory, CompanyServiceSubCategory

    category = CompanyServiceCategory.objects.filter(id=category_id, company=company).first()
    if category is None:
        raise Http404("رسته خدمات یافت نشد.")

    item_count = OrderItemDefinition.objects.filter(company=company, category=category).count()
    subcategory_count = CompanyServiceSubCategory.objects.filter(company=company, category=category).count()
    order_count = Order.objects.filter(company=company, service_category=category).count()

    if request.method == "POST":
        category.delete()
        return redirect(f"/{company.code}/admin/base-data/categories/")

    return render(request, "tenants/admin_base_category_delete.html", {
        "company": company,
        "category": category,
        "item_count": item_count,
        "subcategory_count": subcategory_count,
        "order_count": order_count,
    })



# -----------------------------------------------------------------------------
# Operator management and permissions
# -----------------------------------------------------------------------------

@require_tenant_role("COMPANY_ADMIN")
def admin_operator_list(request: HttpRequest, **kwargs) -> HttpResponse:
    company = request.company
    from django.contrib.auth import get_user_model
    from apps.accounts.models import OperatorPermission
    from apps.accounts.operator_access import (
        get_login_field_name,
        get_operator_queryset,
        get_staff_role_value,
        get_user_display,
        get_user_identifier,
        grouped_permission_items,
        list_operator_permission_items,
        model_has_field,
        set_if_field,
        set_user_display_name,
    )

    User = get_user_model()
    error = ""
    success = ""
    items = list_operator_permission_items()

    if request.method == "POST":
        action = request.POST.get("action", "")

        if action == "create_operator":
            username = (request.POST.get("username") or "").strip().lower()
            phone_raw = (request.POST.get("phone") or "").strip()
            display_name = (request.POST.get("display_name") or "").strip()
            email = (request.POST.get("email") or "").strip()
            password = request.POST.get("password") or "123456"
            is_active = request.POST.get("is_active") == "on"

            if not username:
                error = "نام کاربری اپراتور الزامی است."
            elif User.objects.filter(username=username).exists():
                error = "این نام کاربری قبلاً استفاده شده است. لطفاً نام کاربری دیگری انتخاب کنید."
            else:
                from apps.common.phone_utils import normalize_iran_mobile
                normalized_phone = normalize_iran_mobile(phone_raw) or phone_raw

                operator = User()
                operator.username = username
                operator.phone = normalized_phone
                operator.email = email
                set_user_display_name(operator, display_name)
                if model_has_field(User, "company"):
                    operator.company = company
                if model_has_field(User, "role"):
                    operator.role = get_staff_role_value()
                if model_has_field(User, "is_active"):
                    operator.is_active = is_active
                operator.set_password(password)
                if hasattr(operator, "must_change_password"):
                    operator.must_change_password = (password == "123456")
                operator.save()
                success = "اپراتور ساخته شد. برای تنظیم دسترسی، دکمه ویرایش همان اپراتور را بزنید."

        elif action == "update_operator":
            operator_id = request.POST.get("operator_id")
            operator = get_operator_queryset(company).filter(id=operator_id).first()
            if operator is None:
                error = "اپراتور یافت نشد."
            else:
                display_name = (request.POST.get("display_name") or "").strip()
                email = (request.POST.get("email") or "").strip()
                phone_raw = (request.POST.get("phone") or "").strip()
                is_active = request.POST.get("is_active") == "on"
                selected = set(request.POST.getlist("permissions"))

                set_user_display_name(operator, display_name)
                set_if_field(operator, "email", email)
                if phone_raw:
                    from apps.common.phone_utils import normalize_iran_mobile
                    set_if_field(operator, "phone", normalize_iran_mobile(phone_raw) or phone_raw)
                if model_has_field(User, "is_active"):
                    operator.is_active = is_active
                operator.save()

                for item in items:
                    row, _ = OperatorPermission.objects.get_or_create(
                        company=company,
                        operator=operator,
                        permission_key=item.key,
                        defaults={"is_allowed": False},
                    )
                    row.is_allowed = item.key in selected
                    row.save(update_fields=["is_allowed"])

                success = "اپراتور و دسترسی‌های او ذخیره شد."

        elif action == "change_operator_password":
            operator_id = request.POST.get("operator_id")
            operator = get_operator_queryset(company).filter(id=operator_id).first()
            if operator is None:
                error = "اپراتور یافت نشد."
            else:
                new_password1 = request.POST.get("new_password1") or ""
                new_password2 = request.POST.get("new_password2") or ""

                if not new_password1:
                    error = "رمز جدید نمی‌تواند خالی باشد."
                elif new_password1 != new_password2:
                    error = "رمز جدید و تکرار رمز یکسان نیستند."
                else:
                    operator.set_password(new_password1)
                    operator.save()
                    success = "رمز اپراتور با موفقیت تغییر کرد."

        elif action == "toggle_operator":
            operator_id = request.POST.get("operator_id")
            operator = get_operator_queryset(company).filter(id=operator_id).first()
            if operator is None:
                error = "اپراتور یافت نشد."
            elif model_has_field(User, "is_active"):
                operator.is_active = not operator.is_active
                operator.save(update_fields=["is_active"])
                success = "وضعیت اپراتور تغییر کرد."
            else:
                error = "این مدل کاربر فیلد فعال/غیرفعال ندارد."

        elif action == "delete_operator":
            operator_id = request.POST.get("operator_id")
            operator = get_operator_queryset(company).filter(id=operator_id).first()
            if operator is None:
                error = "اپراتور یافت نشد."
            else:
                OperatorPermission.objects.filter(company=company, operator=operator).delete()
                operator.delete()
                success = "اپراتور حذف شد."

    grouped_master = grouped_permission_items()
    operators = get_operator_queryset(company)

    rows = []
    for operator in operators:
        display = get_user_display(operator)
        identifier = get_user_identifier(operator)
        has_display_name = bool(display and display != identifier)

        allowed_keys = set(
            OperatorPermission.objects.filter(
                company=company,
                operator=operator,
                is_allowed=True,
            ).values_list("permission_key", flat=True)
        )

        grouped = []
        for group, group_items in grouped_master.items():
            grouped.append({
                "group": group,
                "items": [
                    {
                        "key": item.key,
                        "title": item.title,
                        "description": item.description,
                        "action_label": item.action_label,
                        "path_template": item.path_template.replace("{company_code}", company.code),
                        "is_allowed": item.key in allowed_keys,
                    }
                    for item in group_items
                ],
            })

        rows.append({
            "operator": operator,
            "display": display,
            "identifier": identifier,
            "has_display_name": has_display_name,
            "permission_count": len(allowed_keys),
            "grouped_permissions": grouped,
        })

    return render(request, "tenants/admin_operator_list.html", {
        "company": company,
        "operator_rows": rows,
        "error": error,
        "success": success,
    })

# -----------------------------------------------------------------------------
# Order item definitions
# -----------------------------------------------------------------------------


@require_tenant_role("COMPANY_ADMIN")
def admin_base_items(request: HttpRequest, **kwargs) -> HttpResponse:
    company = request.company
    from apps.orders.models import OrderItemDefinition
    from apps.tenants.models import CompanyServiceCategory

    category_id = request.GET.get("category_id") or ""
    categories = CompanyServiceCategory.objects.filter(company=company).order_by("sort_order", "title")
    items = OrderItemDefinition.objects.filter(company=company).select_related("category").order_by(
        "category__sort_order", "category__title", "sort_order", "title",
    )
    if category_id:
        items = items.filter(category_id=category_id)
    return render(request, "tenants/admin_base_items.html", {
        "company": company,
        "items": items,
        "categories": categories,
        "selected_category_id": str(category_id),
    })


@require_tenant_role("COMPANY_ADMIN")
def admin_base_item_create(request: HttpRequest, **kwargs) -> HttpResponse:
    company = request.company
    from apps.orders.models import OrderItemDefinition
    from apps.tenants.models import CompanyServiceCategory

    categories = CompanyServiceCategory.objects.filter(company=company, is_active=True).order_by("sort_order", "title")
    error = ""
    initial_category_id = request.GET.get("category_id") or ""
    if request.method == "POST":
        category = CompanyServiceCategory.objects.filter(
            id=request.POST.get("category_id"), company=company,
        ).first()
        title = request.POST.get("title", "").strip()
        kind = request.POST.get("kind", OrderItemDefinition.Kind.NUMBER)
        if category is None:
            error = "رسته خدمات معتبر انتخاب کنید."
        elif not title:
            error = "عنوان آیتم سفارش الزامی است."
        elif kind not in dict(OrderItemDefinition.Kind.choices):
            error = "نوع آیتم معتبر نیست."
        else:
            item = OrderItemDefinition.objects.create(
                company=company,
                category=category,
                title=title,
                kind=kind,
                sort_order=_parse_positive_int(request.POST.get("sort_order"), 0),
                is_active=bool(request.POST.get("is_active")),
            )
            return redirect(f"/{company.code}/admin/base-data/items/")

    return render(request, "tenants/admin_base_item_form.html", {
        "company": company,
        "item": None,
        "categories": categories,
        "item_kinds": OrderItemDefinition.Kind.choices,
        "initial_category_id": initial_category_id,
        "is_edit": False,
        "error": error,
    })


@require_tenant_role("COMPANY_ADMIN")
def admin_base_item_edit(request: HttpRequest, item_id: int, **kwargs) -> HttpResponse:
    company = request.company
    from apps.orders.models import OrderItemDefinition
    from apps.tenants.models import CompanyServiceCategory

    item = OrderItemDefinition.objects.filter(id=item_id, company=company).select_related("category").first()
    if item is None:
        raise Http404("آیتم سفارش یافت نشد.")

    categories = CompanyServiceCategory.objects.filter(company=company, is_active=True).order_by("sort_order", "title")
    error = ""
    if request.method == "POST":
        category = CompanyServiceCategory.objects.filter(
            id=request.POST.get("category_id"), company=company,
        ).first()
        title = request.POST.get("title", "").strip()
        kind = request.POST.get("kind", item.kind)
        if category is None:
            error = "رسته خدمات معتبر انتخاب کنید."
        elif not title:
            error = "عنوان آیتم سفارش الزامی است."
        elif kind not in dict(OrderItemDefinition.Kind.choices):
            error = "نوع آیتم معتبر نیست."
        else:
            item.category = category
            item.title = title
            item.kind = kind
            item.sort_order = _parse_positive_int(request.POST.get("sort_order"), item.sort_order)
            item.is_active = bool(request.POST.get("is_active"))
            item.save(update_fields=["category", "title", "kind", "sort_order", "is_active", "updated_at"])
            return redirect(f"/{company.code}/admin/base-data/items/")

    return render(request, "tenants/admin_base_item_form.html", {
        "company": company,
        "item": item,
        "categories": categories,
        "item_kinds": OrderItemDefinition.Kind.choices,
        "initial_category_id": item.category_id,
        "is_edit": True,
        "error": error,
    })


@require_tenant_role("COMPANY_ADMIN")
def admin_base_item_delete(request: HttpRequest, item_id: int, **kwargs) -> HttpResponse:
    company = request.company
    from apps.orders.models import OrderItemDefinition

    item = OrderItemDefinition.objects.filter(id=item_id, company=company).select_related("category").first()
    if item is None:
        raise Http404("آیتم سفارش یافت نشد.")

    if request.method == "POST":
        item.is_active = False
        item.save(update_fields=["is_active", "updated_at"])
        return redirect(f"/{company.code}/admin/base-data/items/")

    return render(request, "tenants/admin_base_item_delete.html", {
        "company": company,
        "item": item,
    })


# -----------------------------------------------------------------------------
# Operators
# -----------------------------------------------------------------------------


@require_tenant_role("COMPANY_ADMIN")
def admin_base_operators(request: HttpRequest, **kwargs) -> HttpResponse:
    """List tenant operators/staff users. Editing happens on separate pages."""
    company = request.company
    operators = CompanyUser.objects.filter(
        company=company,
        role__in=[UserRole.COMPANY_ADMIN, UserRole.COMPANY_STAFF],
    ).order_by("role", "first_name", "last_name", "phone")

    return render(request, "tenants/admin_base_operators.html", {
        "company": company,
        "operators": operators,
    })


@require_tenant_role("COMPANY_ADMIN")
def admin_base_operator_create(request: HttpRequest, **kwargs) -> HttpResponse:
    company = request.company
    error = ""
    if request.method == "POST":
        phone = request.POST.get("phone", "").strip()
        if not phone:
            error = "شماره تلفن اپراتور الزامی است."
        elif CompanyUser.objects.filter(phone=phone).exists():
            error = "کاربری با این شماره تلفن وجود دارد."
        else:
            role = request.POST.get("role", UserRole.COMPANY_STAFF)
            if role not in [UserRole.COMPANY_ADMIN, UserRole.COMPANY_STAFF]:
                role = UserRole.COMPANY_STAFF
            _op_password = request.POST.get("password") or "123456"
            operator = CompanyUser.objects.create_user(
                phone=phone,
                password=_op_password,
                company=company,
                role=role,
                first_name=request.POST.get("first_name", "").strip(),
                last_name=request.POST.get("last_name", "").strip(),
                is_active=True,
                must_change_password=(_op_password == "123456"),
            )
            return redirect(f"/{company.code}/admin/settings/operators/{operator.id}/edit/")

    return render(request, "tenants/admin_base_operator_form.html", {
        "company": company,
        "operator": None,
        "operator_roles": _operator_role_choices(),
        "is_edit": False,
        "error": error,
    })


@require_tenant_role("COMPANY_ADMIN")
def admin_base_operator_edit(request: HttpRequest, operator_id: int, **kwargs) -> HttpResponse:
    company = request.company
    operator = CompanyUser.objects.filter(
        id=operator_id,
        company=company,
        role__in=[UserRole.COMPANY_ADMIN, UserRole.COMPANY_STAFF],
    ).first()
    if operator is None:
        raise Http404("اپراتور یافت نشد.")

    error = ""
    if request.method == "POST":
        new_phone = request.POST.get("phone", "").strip()
        if not new_phone:
            error = "شماره تلفن ورود الزامی است."
            return render(request, "tenants/admin_base_operator_form.html", {
                "company": company,
                "operator": operator,
                "operator_roles": _operator_role_choices(),
                "is_edit": True,
                "error": error,
            })
        if CompanyUser.objects.filter(phone=new_phone).exclude(id=operator.id).exists():
            error = "کاربری با این شماره تلفن وجود دارد."
            return render(request, "tenants/admin_base_operator_form.html", {
                "company": company,
                "operator": operator,
                "operator_roles": _operator_role_choices(),
                "is_edit": True,
                "error": error,
            })
        operator.phone = new_phone
        operator.first_name = request.POST.get("first_name", "").strip()
        operator.last_name = request.POST.get("last_name", "").strip()
        role = request.POST.get("role", operator.role)
        if role in [UserRole.COMPANY_ADMIN, UserRole.COMPANY_STAFF]:
            operator.role = role
        operator.is_active = bool(request.POST.get("is_active"))
        new_password = request.POST.get("password", "").strip()
        if new_password:
            operator.set_password(new_password)
        operator.save()
        return redirect(f"/{company.code}/admin/settings/operators/")

    return render(request, "tenants/admin_base_operator_form.html", {
        "company": company,
        "operator": operator,
        "operator_roles": _operator_role_choices(),
        "is_edit": True,
        "error": error,
    })


@require_tenant_role("COMPANY_ADMIN")
def admin_base_operator_delete(request: HttpRequest, operator_id: int, **kwargs) -> HttpResponse:
    company = request.company
    operator = CompanyUser.objects.filter(
        id=operator_id,
        company=company,
        role__in=[UserRole.COMPANY_ADMIN, UserRole.COMPANY_STAFF],
    ).first()
    if operator is None:
        raise Http404("اپراتور یافت نشد.")

    if request.method == "POST":
        operator.is_active = False
        operator.save(update_fields=["is_active"])
        return redirect(f"/{company.code}/admin/settings/operators/")

    return render(request, "tenants/admin_base_operator_delete.html", {
        "company": company,
        "operator": operator,
    })



# =============================================================================
# Phase 25: Cancel Request Admin Review
# =============================================================================


@require_tenant_role("COMPANY_ADMIN", "COMPANY_STAFF")
def admin_cancel_request_approve(request: HttpRequest, order_id: int, **kwargs) -> HttpResponse:
    """Approve a technician's cancel request — order becomes CANCELLED."""
    if request.method != "POST":
        return HttpResponseForbidden("POST only.")

    company = request.company
    order = OrderSelector.get_by_id_for_company(order_id=order_id, company=company)
    if not order:
        raise Http404("سفارش یافت نشد.")

    from apps.orders.cancel_review_service import OrderCancelReviewService

    note = request.POST.get("admin_note", "").strip()
    try:
        OrderCancelReviewService.approve(
            order=order, approved_by=request.user, note=note,
        )
    except ValueError:
        pass  # Silently redirect back — status may have already changed

    return redirect(f"/{company.code}/admin/orders/{order.id}/")


@require_tenant_role("COMPANY_ADMIN", "COMPANY_STAFF")
def admin_cancel_request_reject(request: HttpRequest, order_id: int, **kwargs) -> HttpResponse:
    """Reject a technician's cancel request — restore previous status."""
    if request.method != "POST":
        return HttpResponseForbidden("POST only.")

    company = request.company
    order = OrderSelector.get_by_id_for_company(order_id=order_id, company=company)
    if not order:
        raise Http404("سفارش یافت نشد.")

    from apps.orders.cancel_review_service import OrderCancelReviewService

    note = request.POST.get("admin_note", "").strip()
    try:
        OrderCancelReviewService.reject(
            order=order, rejected_by=request.user, note=note,
        )
    except ValueError:
        pass  # Silently redirect back

    return redirect(f"/{company.code}/admin/orders/{order.id}/")


@require_tenant_role("COMPANY_ADMIN", "COMPANY_STAFF")
def admin_order_return_to_cycle(request: HttpRequest, order_id: int, **kwargs) -> HttpResponse:
    """
    Admin action: بازگشت به چرخه (Return to Cycle).

    Cancels the old order and creates a new cloned order entering the dispatch cycle.
    Permission key: admin_order_return_to_cycle
    """
    if request.method != "POST":
        return HttpResponseForbidden("POST only.")

    company = request.company
    order = OrderSelector.get_by_id_for_company(order_id=order_id, company=company)
    if not order:
        raise Http404("سفارش یافت نشد.")

    from apps.orders.recycle_service import OrderReturnToCycleService

    try:
        new_order = OrderReturnToCycleService.return_to_cycle(
            order=order, performed_by=request.user,
        )
        from django.contrib import messages
        messages.success(
            request,
            f"سفارش #{order.id} لغو شد و سفارش جدید #{new_order.id} وارد چرخه شد.",
        )
        return redirect(f"/{company.code}/admin/orders/{new_order.id}/")
    except ValueError as e:
        from django.contrib import messages
        messages.error(request, str(e))
        return redirect(f"/{company.code}/admin/orders/{order.id}/")


# -----------------------------------------------------------------------------
# Rasti Phase 45B - Robust admin technician actions
# -----------------------------------------------------------------------------

def _rasti45b_model_has_field(model, field_name: str) -> bool:
    try:
        model._meta.get_field(field_name)
        return True
    except Exception:
        return False


def _rasti45b_company_filter_kwargs(model, company):
    for field_name in ("company", "tenant", "business", "organization"):
        if _rasti45b_model_has_field(model, field_name):
            return {field_name: company}
    return {}


def _rasti45b_technician_role_values() -> list[str]:
    values = []
    try:
        from apps.accounts.models import UserRole

        for attr in (
            "TECHNICIAN",
            "COMPANY_TECHNICIAN",
            "SERVICE_TECHNICIAN",
            "COMPANY_STAFF_TECHNICIAN",
        ):
            if hasattr(UserRole, attr):
                values.append(str(getattr(UserRole, attr)))
    except Exception:
        pass

    values.extend(["technician", "company_technician", "service_technician", "tech"])
    return list(dict.fromkeys(values))


def _rasti45b_admin_role_values() -> list[str]:
    values = []
    try:
        from apps.accounts.models import UserRole

        for attr in ("COMPANY_ADMIN", "ADMIN", "OWNER", "PLATFORM_OWNER", "COMPANY_STAFF", "OPERATOR"):
            if hasattr(UserRole, attr):
                values.append(str(getattr(UserRole, attr)))
    except Exception:
        pass

    values.extend(["company_admin", "admin", "owner", "platform_owner", "company_staff", "operator"])
    return list(dict.fromkeys(values))


def _rasti45b_get_related_user(obj):
    from django.contrib.auth import get_user_model

    User = get_user_model()
    if isinstance(obj, User):
        return obj

    for attr in ("user", "account", "member", "profile_user"):
        try:
            value = getattr(obj, attr, None)
            if value is not None and isinstance(value, User):
                return value
        except Exception:
            pass

    try:
        for field in obj._meta.fields:
            try:
                if getattr(field, "remote_field", None) and field.remote_field.model == User:
                    value = getattr(obj, field.name, None)
                    if value is not None:
                        return value
            except Exception:
                pass
    except Exception:
        pass

    return None


def _rasti45b_is_safe_technician_user(user) -> bool:
    if user is None:
        return False

    role = str(getattr(user, "role", "") or "")
    if role in _rasti45b_admin_role_values():
        return False

    if role in _rasti45b_technician_role_values():
        return True

    # If there is no role field/value, treat non-staff non-superuser as safe technician-like.
    if getattr(user, "is_superuser", False):
        return False
    if getattr(user, "is_staff", False):
        return False

    return True


def _rasti45b_find_technician_target(company, technician_id):
    from django.apps import apps
    from django.contrib.auth import get_user_model

    User = get_user_model()

    # 1) Try User directly.
    qs = User.objects.all()
    if _rasti45b_model_has_field(User, "company"):
        qs = qs.filter(company=company)

    user = qs.filter(id=technician_id).first()
    if user is not None and _rasti45b_is_safe_technician_user(user):
        return user, user

    # 2) Try models that look like technician/service-person profiles.
    candidate_models = []
    for model in apps.get_models():
        model_name = model.__name__.lower()
        app_label = getattr(model._meta, "app_label", "").lower()
        if (
            "technician" in model_name
            or "technician" in app_label
            or model_name in {"serviceprovider", "serviceperson", "worker", "staffprofile"}
        ):
            candidate_models.append(model)

    for model in candidate_models:
        try:
            filters = {"id": technician_id}
            filters.update(_rasti45b_company_filter_kwargs(model, company))
            obj = model.objects.filter(**filters).first()
        except Exception:
            obj = None

        if obj is None:
            continue

        related_user = _rasti45b_get_related_user(obj)
        if related_user is None or _rasti45b_is_safe_technician_user(related_user):
            return obj, related_user

    # 3) Last fallback: any non-admin user by id in company.
    if user is not None:
        return user, user

    return None, None


def _rasti45b_toggle_active_on(obj, user):
    changed = False

    for target in (obj, user):
        if target is None:
            continue
        model = target.__class__

        for field_name in ("is_active", "active", "is_enabled", "enabled"):
            if _rasti45b_model_has_field(model, field_name):
                current = bool(getattr(target, field_name))
                setattr(target, field_name, not current)
                try:
                    target.save(update_fields=[field_name])
                except Exception:
                    target.save()
                changed = True
                break

    if not changed and obj is not None:
        obj.save()


def _rasti45b_delete_target(obj, user):
    from django.contrib.auth import get_user_model

    User = get_user_model()

    if obj is None and user is None:
        return

    if obj is not None and not isinstance(obj, User):
        obj.delete()
        if user is not None and _rasti45b_is_safe_technician_user(user):
            try:
                user.delete()
            except Exception:
                pass
        return

    if user is not None and _rasti45b_is_safe_technician_user(user):
        user.delete()


@require_tenant_role("COMPANY_ADMIN")
def admin_technician_toggle_active(request: HttpRequest, technician_id: int, **kwargs) -> HttpResponse:
    company = request.company
    obj, user = _rasti45b_find_technician_target(company, technician_id)

    if obj is None and user is None:
        raise Http404("نیروی خدماتی یافت نشد.")

    _rasti45b_toggle_active_on(obj, user)
    return redirect(f"/{company.code}/admin/technicians/")


@require_tenant_role("COMPANY_ADMIN")
def admin_technician_delete(request: HttpRequest, technician_id: int, **kwargs) -> HttpResponse:
    company = request.company
    obj, user = _rasti45b_find_technician_target(company, technician_id)

    if obj is None and user is None:
        raise Http404("نیروی خدماتی یافت نشد.")

    _rasti45b_delete_target(obj, user)
    return redirect(f"/{company.code}/admin/technicians/")


# -----------------------------------------------------------------------------
# Rasti Phase 46 - Jalali admin orders page
# -----------------------------------------------------------------------------

def _rasti46_model_has_field(model, field_name: str) -> bool:
    try:
        model._meta.get_field(field_name)
        return True
    except Exception:
        return False


def _rasti46_get_field(model, candidates):
    for name in candidates:
        try:
            return model._meta.get_field(name)
        except Exception:
            continue
    return None


def _rasti46_get_order_model():
    from django.apps import apps

    for app_label, model_name in (
        ("orders", "Order"),
        ("orders", "ServiceOrder"),
        ("orders", "CustomerOrder"),
    ):
        try:
            model = apps.get_model(app_label, model_name)
            if model is not None:
                return model
        except Exception:
            pass

    for model in apps.get_models():
        if model.__name__.lower() == "order":
            return model

    return None


def _rasti46_get_company_filter(model, company):
    for field_name in ("company", "tenant", "business", "organization"):
        if _rasti46_model_has_field(model, field_name):
            return {field_name: company}
    return {}


def _rasti46_text(value, empty="-"):
    if value in (None, ""):
        return empty
    text = str(value)
    text = text.replace("Technician:", "نیروی خدماتی:")
    text = text.replace(":Technician", ":نیروی خدماتی")
    return text


def _rasti46_short(value, length=60):
    text = _rasti46_text(value, "")
    if not text:
        return "-"
    return text if len(text) <= length else text[:length] + "..."


def _rasti46_get_value(obj, candidates, short=False):
    for candidate in candidates:
        try:
            if "__" in candidate:
                value = obj
                for part in candidate.split("__"):
                    value = getattr(value, part)
                    if value is None:
                        break
            else:
                value = getattr(obj, candidate)
            if callable(value):
                value = value()
            if value not in (None, ""):
                return _rasti46_short(value) if short else _rasti46_text(value)
        except Exception:
            continue
    return "-"


def _rasti46_status_label_from_choices(model, field_name, raw):
    if raw in (None, ""):
        return "-"
    try:
        from apps.common.templatetags.fa_labels import status_fa
    except Exception:
        status_fa = lambda value: str(value)

    try:
        field = model._meta.get_field(field_name)
        choices = getattr(field, "choices", None) or []
        choice_map = {str(value): str(label) for value, label in choices}
        label = choice_map.get(str(raw), str(raw))
        translated = status_fa(raw)
        return translated if translated != str(raw) else status_fa(label)
    except Exception:
        return status_fa(raw)


def _rasti46_get_status_value(obj, candidates):
    for candidate in candidates:
        try:
            raw = getattr(obj, candidate)
            if raw in (None, ""):
                continue
            return _rasti46_status_label_from_choices(obj.__class__, candidate, raw)
        except Exception:
            continue
    return "-"


def _rasti46_get_date_value(obj, candidates, include_time=False):
    from apps.common.jalali import format_jalali, format_jalali_datetime

    for candidate in candidates:
        try:
            value = getattr(obj, candidate)
            if value not in (None, ""):
                return format_jalali_datetime(value) if include_time else format_jalali(value)
        except Exception:
            continue
    return "-"


def _rasti46_build_order_url(company, order):
    return f"/{company.code}/admin/orders/{order.pk}/"


def _rasti46_column_definitions():
    return [
        {
            "key": "order_number",
            "title": "شماره سفارش",
            "candidates": ["order_number", "number", "code", "tracking_code", "id"],
            "type": "text",
        },
        {
            "key": "service_category",
            "title": "رسته خدمات",
            "candidates": ["service_category", "category", "service__category", "service"],
            "type": "text",
        },
        {
            "key": "customer_name",
            "title": "نام مشتری",
            "candidates": ["customer_name", "name", "customer__full_name", "customer__name", "customer"],
            "type": "text",
        },
        {
            "key": "customer_phone",
            "title": "شماره تماس",
            "candidates": ["customer_phone", "phone", "mobile", "customer__phone", "customer__mobile"],
            "type": "text",
        },
        {
            "key": "address_short",
            "title": "آدرس کوتاه",
            "candidates": ["address_short", "short_address", "address", "customer_address"],
            "type": "short",
        },
        {
            "key": "address_full",
            "title": "آدرس کامل",
            "candidates": ["address_full", "full_address", "address", "customer_address"],
            "type": "text",
        },
        {
            "key": "description_short",
            "title": "توضیحات کوتاه",
            "candidates": ["description_short", "short_description", "description", "note", "customer_note"],
            "type": "short",
        },
        {
            "key": "description_full",
            "title": "توضیحات کامل",
            "candidates": ["description_full", "full_description", "description", "note", "customer_note"],
            "type": "text",
        },
        {
            "key": "status",
            "title": "وضعیت",
            "candidates": ["status", "state", "order_status"],
            "type": "status",
        },
        {
            "key": "technician",
            "title": "نیروی خدماتی",
            "candidates": ["technician", "assigned_technician", "accepted_by", "assignee", "worker", "service_provider"],
            "type": "text",
        },
        {
            "key": "scheduled_date",
            "title": "تاریخ انجام",
            "candidates": ["scheduled_date", "service_date", "done_at", "completed_at", "appointment_date"],
            "type": "date",
        },
        {
            "key": "created_at",
            "title": "تاریخ ثبت",
            "candidates": ["created_at", "created", "inserted_at", "date_joined"],
            "type": "datetime",
        },
        {
            "key": "accepted_at",
            "title": "تاریخ پذیرش",
            "candidates": ["accepted_at", "accepted_date", "assigned_at"],
            "type": "datetime",
        },
        {
            "key": "survey_rating",
            "title": "امتیاز نظرسنجی",
            "candidates": ["survey_rating", "rating", "review_rating", "feedback_rating"],
            "type": "text",
        },
        {
            "key": "survey_comment",
            "title": "توضیح نظرسنجی",
            "candidates": ["survey_comment", "review_comment", "feedback_comment", "survey_text"],
            "type": "text",
        },
        {
            "key": "survey_received_at",
            "title": "زمان دریافت نظرسنجی",
            "candidates": ["survey_received_at", "reviewed_at", "feedback_at", "survey_at"],
            "type": "datetime",
        },
        {
            "key": "actions",
            "title": "عملیات",
            "candidates": [],
            "type": "actions",
        },
    ]


def _rasti46_render_cell(order, company, col):
    col_type = col["type"]
    if col_type == "actions":
        return ""
    if col_type == "status":
        return _rasti46_get_status_value(order, col["candidates"])
    if col_type == "date":
        return _rasti46_get_date_value(order, col["candidates"], include_time=False)
    if col_type == "datetime":
        return _rasti46_get_date_value(order, col["candidates"], include_time=True)
    if col_type == "short":
        return _rasti46_get_value(order, col["candidates"], short=True)
    return _rasti46_get_value(order, col["candidates"], short=False)


def _rasti46_export_xlsx_or_csv(rows, selected_columns):
    from django.http import HttpResponse

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "orders"
        ws.sheet_view.rightToLeft = True

        headers = [col["title"] for col in selected_columns if col["key"] != "actions"]
        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        for row in rows:
            ws.append([value for value, col in zip(row["values"], selected_columns) if col["key"] != "actions"])

        for idx, _ in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = 22

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="orders.xlsx"'
        wb.save(response)
        return response

    except Exception:
        import csv

        response = HttpResponse(content_type="text/csv; charset=utf-8-sig")
        response["Content-Disposition"] = 'attachment; filename="orders.csv"'
        response.write("\ufeff")
        writer = csv.writer(response)
        writer.writerow([col["title"] for col in selected_columns if col["key"] != "actions"])
        for row in rows:
            writer.writerow([value for value, col in zip(row["values"], selected_columns) if col["key"] != "actions"])
        return response


def _rasti46_get_technician_field(Order):
    return _rasti46_get_field(Order, ["technician", "assigned_technician", "accepted_by", "assignee", "worker", "service_provider"])


def _rasti46_get_status_field(Order):
    return _rasti46_get_field(Order, ["status", "state", "order_status"])


def _rasti46_get_created_field(Order):
    return _rasti46_get_field(Order, ["created_at", "created", "inserted_at", "date_joined"])


def _rasti46_distinct_statuses(qs, status_field):
    if not status_field:
        return []

    try:
        values = qs.exclude(**{f"{status_field.name}__isnull": True}).values_list(status_field.name, flat=True).distinct()
        result = []
        model = getattr(qs, "model", None)
        for value in values:
            if value not in (None, ""):
                raw = str(value)
                label = _rasti46_status_label_from_choices(model, status_field.name, raw) if model else raw
                result.append({"value": raw, "label": label})
        return sorted(result, key=lambda item: item["label"])
    except Exception:
        return []


def _rasti46_technician_options(qs, technician_field):
    if not technician_field:
        return []

    result = []

    try:
        if getattr(technician_field, "remote_field", None) and technician_field.remote_field:
            model = technician_field.remote_field.model
            ids = qs.exclude(**{f"{technician_field.name}__isnull": True}).values_list(
                f"{technician_field.name}__id", flat=True
            ).distinct()
            for obj in model.objects.filter(id__in=list(ids)):
                result.append({"id": str(obj.pk), "title": str(obj)})
        else:
            values = qs.exclude(**{f"{technician_field.name}__isnull": True}).values_list(
                technician_field.name, flat=True
            ).distinct()
            for value in values:
                if value not in (None, ""):
                    result.append({"id": str(value), "title": str(value)})
    except Exception:
        pass

    return sorted(result, key=lambda item: item["title"])


@require_tenant_role("COMPANY_ADMIN", "COMPANY_STAFF")
def admin_orders(request, **kwargs):
    from django.db.models import Q
    from django.shortcuts import render
    from django.utils import timezone
    from datetime import datetime, time
    from apps.common.jalali import parse_jalali_date

    company = getattr(request, "company", None)
    if company is None:
        from django.http import Http404
        raise Http404("Company not found.")

    Order = _rasti46_get_order_model()
    if Order is None:
        from django.http import Http404
        raise Http404("Order model not found.")

    base_qs = Order.objects.all()
    company_filter = _rasti46_get_company_filter(Order, company)
    if company_filter:
        base_qs = base_qs.filter(**company_filter)

    qs = base_qs

    q = (request.GET.get("q") or "").strip()
    from_date_raw = (request.GET.get("from_date") or "").strip()
    to_date_raw = (request.GET.get("to_date") or "").strip()
    technician_id = (request.GET.get("technician") or "").strip()
    status = (request.GET.get("status") or "").strip()

    status_field = _rasti46_get_status_field(Order)
    technician_field = _rasti46_get_technician_field(Order)
    created_field = _rasti46_get_created_field(Order)

    if q:
        query = Q()
        for field in Order._meta.fields:
            if field.get_internal_type() in ("CharField", "TextField", "EmailField", "SlugField"):
                query |= Q(**{f"{field.name}__icontains": q})
        if query:
            qs = qs.filter(query)

    if status and status_field:
        qs = qs.filter(**{status_field.name: status})

    if technician_id and technician_field:
        try:
            if getattr(technician_field, "remote_field", None) and technician_field.remote_field:
                qs = qs.filter(**{f"{technician_field.name}__id": technician_id})
            else:
                qs = qs.filter(**{technician_field.name: technician_id})
        except Exception:
            pass

    if created_field:
        start_date = parse_jalali_date(from_date_raw)
        end_date = parse_jalali_date(to_date_raw)

        if start_date:
            if created_field.get_internal_type() == "DateTimeField":
                start_dt = timezone.make_aware(datetime.combine(start_date, time.min)) if timezone.is_naive(datetime.combine(start_date, time.min)) else datetime.combine(start_date, time.min)
                qs = qs.filter(**{f"{created_field.name}__gte": start_dt})
            else:
                qs = qs.filter(**{f"{created_field.name}__gte": start_date})

        if end_date:
            if created_field.get_internal_type() == "DateTimeField":
                end_dt = timezone.make_aware(datetime.combine(end_date, time.max)) if timezone.is_naive(datetime.combine(end_date, time.max)) else datetime.combine(end_date, time.max)
                qs = qs.filter(**{f"{created_field.name}__lte": end_dt})
            else:
                qs = qs.filter(**{f"{created_field.name}__lte": end_date})

    order_by_field = created_field.name if created_field else "id"
    try:
        qs = qs.order_by(f"-{order_by_field}")
    except Exception:
        qs = qs.order_by("-id")

    columns = _rasti46_column_definitions()
    selected_keys = request.GET.getlist("columns")
    if not selected_keys:
        # Default visible columns — most important for daily operations
        selected_keys = [
            "order_number",
            "service_category",
            "customer_name",
            "customer_phone",
            "status",
            "technician",
            "created_at",
            "actions",
        ]

    selected_columns = [col for col in columns if col["key"] in selected_keys]

    orders = list(qs[:1000])
    rows = []
    for order in orders:
        values = [_rasti46_render_cell(order, company, col) for col in selected_columns]
        rows.append({
            "order": order,
            "values": values,
            "detail_url": _rasti46_build_order_url(company, order),
            "edit_url": f"/{company.code}/admin/orders/{order.pk}/edit/",
            "assign_url": f"/{company.code}/admin/orders/{order.pk}/assign/",
            "status_raw": str(getattr(order, "status", "") or ""),
        })

    if request.GET.get("export") == "excel":
        return _rasti46_export_xlsx_or_csv(rows, selected_columns)

    # Build per-user permission flags for template action visibility
    from apps.accounts.operator_access import is_company_admin as _is_admin, operator_has_permission
    user = request.user
    user_is_admin = _is_admin(user)
    can_edit = user_is_admin or operator_has_permission(company=company, operator=user, permission_key="admin_order_edit")
    can_create = user_is_admin or operator_has_permission(company=company, operator=user, permission_key="admin_order_create")
    can_assign = user_is_admin or operator_has_permission(company=company, operator=user, permission_key="admin_order_assign")

    # Track which column indices are status type for badge rendering in template
    status_col_indices = [i for i, col in enumerate(selected_columns) if col["type"] == "status"]

    context = {
        "company": company,
        "orders": orders,
        "rows": rows,
        "columns": columns,
        "selected_columns": selected_columns,
        "selected_keys": selected_keys,
        "status_col_indices": status_col_indices,
        "q": q,
        "from_date": from_date_raw,
        "to_date": to_date_raw,
        "selected_technician": technician_id,
        "selected_status": status,
        "status_options": _rasti46_distinct_statuses(base_qs, status_field),
        "technician_options": _rasti46_technician_options(base_qs, technician_field),
        "result_count": qs.count() if hasattr(qs, "count") else len(orders),
        "user_is_admin": user_is_admin,
        "can_edit": can_edit,
        "can_create": can_create,
        "can_assign": can_assign,
    }

    return render(request, "tenants/admin_orders.html", context)

